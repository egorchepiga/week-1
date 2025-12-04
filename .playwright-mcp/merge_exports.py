#!/usr/bin/env python3
"""
Merge all app-database.com export files into one combined Excel file.
Extracts hyperlinks from Title column into separate URL column.
Source files are NOT deleted.
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import warnings

# Suppress DrawingML warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Configuration
SOURCE_DIR = Path(__file__).parent
OUTPUT_FILE = SOURCE_DIR / f"app-database-COMBINED-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
HEADER_ROW = 8  # 1-indexed for openpyxl (row 8 = headers)
DATA_START_ROW = 9  # 1-indexed for openpyxl
TITLE_COL = 2  # Column B = Title

# Column names in order
COLUMNS = [
    'Logo', 'Title', 'Title_URL', 'Users', 'ID', 'Version', 'Rating', 'Reviews',
    'Manifest', 'Lang', 'Categories', 'Featured', 'Trader', 'Website',
    'Email', 'Size', 'Updated at', 'Added at'
]

def get_export_files():
    """Find all export xlsx files and sort them by page number."""
    patterns = [
        "app_database_com-cws-export-*.xlsx",
        "app-database-com-cws-export-*.xlsx",
    ]

    files = []
    for pattern in patterns:
        files.extend(SOURCE_DIR.glob(pattern))

    # Filter out the combined file
    files = [f for f in files if "COMBINED" not in f.name]

    # Sort by page number
    def get_page_num(filepath):
        name = filepath.stem
        try:
            return int(name.split('_')[-1])
        except ValueError:
            return 0

    return sorted(files, key=get_page_num)

def read_export_file_with_hyperlinks(filepath):
    """Read export file and extract hyperlinks from Title column."""
    print(f"  Reading: {filepath.name}")

    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    # Find the last row with data
    max_row = ws.max_row

    # Read data starting from DATA_START_ROW
    data = []
    for row_num in range(DATA_START_ROW, max_row + 1):
        row_data = []

        for col_num in range(1, 18):  # 17 original columns
            cell = ws.cell(row=row_num, column=col_num)
            value = cell.value

            # After Title (col 2), insert Title_URL
            if col_num == TITLE_COL:
                row_data.append(value)  # Title text
                # Extract hyperlink URL
                if cell.hyperlink and cell.hyperlink.target:
                    row_data.append(cell.hyperlink.target)
                else:
                    row_data.append(None)
            else:
                row_data.append(value)

        # Skip empty rows
        if any(v is not None for v in row_data):
            data.append(row_data)

    wb.close()

    df = pd.DataFrame(data, columns=COLUMNS)
    print(f"    -> {len(df)} rows")
    return df

def main():
    print("=" * 60)
    print("Merging app-database.com export files")
    print("(with hyperlink extraction from Title column)")
    print("=" * 60)

    files = get_export_files()
    print(f"\nFound {len(files)} files to merge:\n")

    if not files:
        print("ERROR: No export files found!")
        return

    # Read and combine all files
    dataframes = []
    for filepath in files:
        df = read_export_file_with_hyperlinks(filepath)
        dataframes.append(df)

    # Concatenate all dataframes
    print("\nCombining all dataframes...")
    combined = pd.concat(dataframes, ignore_index=True)

    # Remove duplicates by ID
    initial_count = len(combined)
    combined = combined.drop_duplicates(subset=['ID'], keep='first')
    duplicates_removed = initial_count - len(combined)

    print(f"\nResult:")
    print(f"  Total rows: {len(combined)}")
    print(f"  Duplicates removed: {duplicates_removed}")
    print(f"  Columns: {len(combined.columns)}")
    print(f"  Column list: {list(combined.columns)}")

    # Save to Excel
    print(f"\nSaving to: {OUTPUT_FILE.name}")
    combined.to_excel(OUTPUT_FILE, index=False, engine='openpyxl')

    print(f"\nDone! Output file size: {OUTPUT_FILE.stat().st_size / 1024 / 1024:.2f} MB")
    print("=" * 60)

if __name__ == "__main__":
    main()
